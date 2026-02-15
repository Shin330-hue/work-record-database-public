#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
図面データ入力用Excelテンプレート作成スクリプト
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_excel_template():
    """図面データ入力用のExcelテンプレートを作成"""
    
    # 出力ファイルパス
    output_file = "../doc/図面データ入力テンプレート.xlsx"
    
    # ExcelWriterオブジェクトを作成
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # シート1: 基本情報（必須）
        basic_info_data = {
            '項目': ['図面番号', '会社ID', '会社名', '会社短縮名', '製品ID', '製品名', '製品カテゴリ', '図面タイトル'],
            '必須/任意': ['必須', '必須', '必須', '必須', '必須', '必須', '必須', '必須'],
            '値': ['0D127100014', 'chuo-tekko', '有限会社中央鉄工所', '中央鉄工所', 'precision-parts', 'チェーンソー', 'ブラケット', 'ブラケット（チェーンソー）加工手順'],
            '説明': ['図面の一意識別番号', '会社の一意識別子', '会社の正式名称', '会社の短縮名', '製品の一意識別子', '製品の名称', '製品の分類', '図面のタイトル']
        }
        basic_info_df = pd.DataFrame(basic_info_data)
        basic_info_df.to_excel(writer, sheet_name='基本情報（必須）', index=False)
        
        # シート2: 検索・分類情報（必須）
        search_info_data = {
            '項目': ['キーワード', '難易度', '推定時間', '機械タイプ', '画像有無', '動画有無', '図面有無'],
            '必須/任意': ['必須', '必須', '必須', '必須', '必須', '必須', '必須'],
            '値': ['ブラケット,チェーンソー,精密,加工,マシニング', '中級', '500分', 'マシニングセンタ', 'あり', 'あり', 'あり'],
            '説明': ['カンマ区切り', '初級/中級/上級', '総作業時間', '使用機械', 'あり/なし', 'あり/なし', 'あり/なし']
        }
        search_info_df = pd.DataFrame(search_info_data)
        search_info_df.to_excel(writer, sheet_name='検索・分類情報（必須）', index=False)
        
        # シート3: 作業手順概要
        overview_data = {
            '項目': ['作業説明', '警告事項1', '警告事項2', '警告事項3', '準備時間', '加工時間', '必要工具'],
            '必須/任意': ['必須', '任意', '任意', '任意', '必須', '必須', '必須'],
            '値': [
                '産業機械の骨格となるメインフレーム部品の加工を行います。SS400材からマシニングセンタで外形・ポケット加工を行い、ラジアルボール盤で精密な穴あけ・タップ加工を実施します。寸法精度と表面粗さに注意が必要な重要部品です。',
                '材料のひずみに注意し、十分な除去加工を行ってください',
                '穴位置の精度が組み立て精度に直結するため、慎重な段取りが必要です',
                '切削油を十分に供給し、工具寿命の延長を図ってください',
                '45分',
                '135分',
                'φ20エンドミル,φ12エンドミル,φ8ドリル,φ6ドリル,M8タップ,M6タップ'
            ],
            '説明': ['作業の概要説明', '重要な注意事項', '重要な注意事項', '重要な注意事項', '準備にかかる時間', '実際の加工時間', 'カンマ区切り']
        }
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name='作業手順概要', index=False)
        
        # シート4: 作業ステップ
        work_steps_data = {
            'ステップ番号': [1, 2, 3],
            'タイトル': [
                'マシニングセンタでの外形・ポケット加工',
                'ラジアルボール盤での穴あけ加工',
                'タップ加工・最終検査'
            ],
            '説明': [
                'SS400材からメインフレームの外形形状とポケット部の荒加工・仕上げ加工を行います',
                '組み立て用の取付穴とボルト穴をラジアルボール盤で精密に加工します',
                'ボルト用ねじ穴の加工と最終的な寸法・品質検査を実施します'
            ],
            '詳細手順': [
                '材料寸法確認（300×200×50mm SS400）および外観検査を実施;マシニングセンタのバイスに材料をセット、ダイヤルゲージで水平出し確認;φ20エンドミルで外形荒加工（切り込み3mm、送り500mm/min、回転数800rpm）',
                '加工済みワークをラジアルボール盤の定盤に設置、ストレートエッジで基準面確認;図面に基づき穴位置をケガキ、ポンチングで穴位置マーキング;φ8ドリルで下穴加工（8箇所、貫通穴、回転数600rpm、送り0.15mm/rev）',
                'φ8穴にM8×1.25タップ加工（8箇所、タップ回転数150rpm、切削油使用）;φ6穴にM6×1.0タップ加工（12箇所、タップ回転数180rpm、切削油使用）;ねじゲージによるねじ精度確認（6H級）'
            ],
            '時間': ['90分', '45分', '45分'],
            '警告レベル': ['important', 'caution', 'critical'],
            '画像ファイル': [
                'step01-material-setup.jpg,step01-machining-roughing.jpg,step01-pocket-finishing.jpg',
                'step02-drilling-setup.jpg,step02-hole-positioning.jpg,step02-drilling-process.jpg',
                'step03-tapping.jpg,step03-inspection.jpg,step03-final-product.jpg'
            ],
            '動画ファイル': ['step01-machining-process.mp4', '', 'step03-final-inspection.mp4']
        }
        work_steps_df = pd.DataFrame(work_steps_data)
        work_steps_df.to_excel(writer, sheet_name='作業ステップ', index=False)
        
        # シート5: 切削条件
        cutting_conditions_data = {
            'ステップ番号': [1, 1, 2, 2, 3, 3],
            '加工タイプ': ['荒加工', '仕上げ加工', '穴あけ8mm', '穴あけ6mm', 'タップM8', 'タップM6'],
            '工具': [
                'φ20 4枚刃エンドミル（TiAlNコーティング）',
                'φ12 4枚刃エンドミル（TiCNコーティング）',
                'φ8 ハイスドリル（ストレートシャンク）',
                'φ6 ハイスドリル（ストレートシャンク）',
                'M8×1.25 ハイスタップ（TiNコーティング）',
                'M6×1.0 ハイスタップ（TiNコーティング）'
            ],
            '回転数': ['800rpm', '1200rpm', '600rpm', '800rpm', '150rpm', '180rpm'],
            '送り速度': ['500mm/min', '300mm/min', '0.15mm/rev', '0.12mm/rev', '187.5mm/min', '180mm/min'],
            '切り込み量': ['3.0mm', '0.5mm', '貫通', '30mm', '自動送り', '自動送り'],
            'ステップオーバー': ['12.0mm', '8.0mm', '', '', '', ''],
            '切削油': ['水溶性切削油（7%）', '水溶性切削油（7%）', '切削油（ストレート油）', '切削油（ストレート油）', 'タッピング油', 'タッピング油']
        }
        cutting_conditions_df = pd.DataFrame(cutting_conditions_data)
        cutting_conditions_df.to_excel(writer, sheet_name='切削条件', index=False)
        
        # シート6: 品質チェック
        quality_check_data = {
            'ステップ番号': [1, 1, 1, 1, 2, 2, 2, 2, 3, 3, 3, 3, 3],
            'チェック項目': [
                '外形寸法（±0.1mm）', 'ポケット深さ（±0.05mm）', '表面粗さ（Ra3.2以下）', '直角度（0.05mm以下）',
                '穴径（φ8 +0/+0.05mm）', '穴径（φ6 +0/+0.03mm）', '穴位置度（±0.03mm）', '穴の真円度（0.01mm以下）',
                'ねじ精度（M8×1.25-6H）', 'ねじ精度（M6×1.0-6H）', '外形寸法（図面指示±0.1mm）', 'ポケット寸法（図面指示±0.05mm）', '表面粗さ（Ra3.2以下）'
            ],
            '測定工具': [
                'ノギス', 'ハイトゲージ', '表面粗さ計', 'スコヤ',
                'プラグゲージ', 'プラグゲージ', '座標測定機', '真円度測定機',
                'ねじゲージ', 'ねじゲージ', 'ノギス', 'ハイトゲージ', '表面粗さ計'
            ],
            '公差': ['±0.1mm', '±0.05mm', 'Ra3.2以下', '0.05mm以下', '+0/+0.05mm', '+0/+0.03mm', '±0.03mm', '0.01mm以下', '6H級', '6H級', '±0.1mm', '±0.05mm', 'Ra3.2以下']
        }
        quality_check_df = pd.DataFrame(quality_check_data)
        quality_check_df.to_excel(writer, sheet_name='品質チェック', index=False)
        
        # シート7: トラブルシューティング
        troubleshooting_data = {
            '問題': ['外形寸法不良', '穴位置精度不良', 'タップ折れ', '表面粗さ不良'],
            '原因': ['工具摩耗または機械の熱変位', 'ケガキ不正確またはドリル逃げ', '切削油不足または無理な送り', '切削条件不適切または工具状態不良'],
            '解決方法': ['工具交換と十分な暖機運転を実施', 'ケガキ再確認、ドリル状態チェック、段取り見直し', '切削油十分供給、送り速度調整、タップ状態確認', '切削条件見直し、工具交換、切削油見直し']
        }
        troubleshooting_df = pd.DataFrame(troubleshooting_data)
        troubleshooting_df.to_excel(writer, sheet_name='トラブルシューティング', index=False)
        
        # シート8: 関連情報
        related_info_data = {
            '項目': ['関連図面1', '関連図面2', '関連アイデア1', '関連アイデア2'],
            '必須/任意': ['任意', '任意', '任意', '任意'],
            '値': ['FR2024002138492', 'BR2024001345671', 'thin-wall/thin-wall_001', 'thin-wall/thin-wall_002'],
            '説明': ['類似フレーム', '組み立て部品', '関連する加工アイデア', '関連する加工アイデア']
        }
        related_info_df = pd.DataFrame(related_info_data)
        related_info_df.to_excel(writer, sheet_name='関連情報', index=False)
        
        # シート9: 改訂履歴
        revision_history_data = {
            'バージョン': ['1.0', '1.1', '1.2', '1.3'],
            '日付': ['2024-02-15', '2024-05-10', '2024-08-25', '2024-11-20'],
            '作成者': ['田中工場長', '佐藤主任', '山田技師', '田中工場長'],
            '変更内容': ['初版作成', '切削条件を最適化、穴あけ精度向上', '品質チェック項目追加、トラブルシューティング強化', 'タップ加工条件見直し、最終検査手順改良']
        }
        revision_history_df = pd.DataFrame(revision_history_data)
        revision_history_df.to_excel(writer, sheet_name='改訂履歴', index=False)
        
        # シート10: 使用説明
        instructions_data = {
            '項目': [
                '必須項目の識別',
                '任意項目の扱い',
                '複数値の入力方法',
                'テンプレートの使用方法',
                'JSON変換の依頼'
            ],
            '説明': [
                '必須項目は黄色でハイライトされています。必ず入力してください。',
                '任意項目は空欄でも構いません。',
                '複数の値を入力する場合は、カンマまたはセミコロンで区切ってください。',
                'このテンプレートをコピーして新しい図面用に使用してください。',
                '入力完了後、LLMにJSON変換を依頼してください。'
            ]
        }
        instructions_df = pd.DataFrame(instructions_data)
        instructions_df.to_excel(writer, sheet_name='使用説明', index=False)
    
    # スタイルを適用
    apply_styles(output_file)
    
    print(f"Excelテンプレートが作成されました: {output_file}")

def apply_styles(file_path):
    """Excelファイルにスタイルを適用"""
    wb = openpyxl.load_workbook(file_path)
    
    # 黄色のフィル
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # 太字フォント
    bold_font = Font(bold=True)
    # 中央揃え
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # ヘッダー行を太字に
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = center_alignment
        
        # 必須項目の行を黄色にハイライト
        for row in ws.iter_rows(min_row=2):
            if len(row) > 1 and row[1].value == '必須':
                for cell in row:
                    cell.fill = yellow_fill
    
    wb.save(file_path)

if __name__ == "__main__":
    create_excel_template() 